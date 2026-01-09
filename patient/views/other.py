import os

from django.http import HttpResponseRedirect, JsonResponse
from django.contrib import messages
from django.views.generic import CreateView, UpdateView, DetailView, ListView, TemplateView, View
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from patient.models import Patient, CornealTopography, ACCustomization, ReviewResult, BasicParams, RelationshipTable
from patient.forms import (
    ReviewResultForm,
    ParamsModifyForm,
)
from django.conf import settings

from services.zs_tear_film import TearFilmHeightCalculator, FluorescentStaining
from services.z_leimo import TEARFILMDATA
from services.z_qcode import txt_to_qrcode

from patient.views.constants import *

from django.urls import reverse
from django.http import HttpResponseRedirect

import csv  # <-- 1. 引入 Python 的 csv 模块
from django.http import HttpResponse  # <-- 2. 引入 HttpResponse

from django.views import View
from django.views.decorators.csrf import csrf_exempt
import io

import random
import string
import numpy as np
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment
import base64

# =============   ↓↓↓ 2. AI码 生成逻辑 ↓↓↓   ================
# ==========================================================


def sum_digits_from_float(value):
    """
    计算一个浮点数所有数字的和。
    例如: -1.234 -> 1+2+3+4 = 10
    """
    try:
        # 1. 取绝对值，转为字符串
        # 2. re.sub(...) 会移除所有非数字的字符 (包括小数点和负号)
        val_str = re.sub(r'[^0-9]', '', str(abs(value)))
        if not val_str:
            return 0
        # 3. 将每个数字相加
        return sum(int(digit) for digit in val_str)
    except Exception:
        return 0  # 发生错误时返回 0

def map_remainder_to_char(remainder):
    """
    根据规则将 1-36 的余数映射到字符
    - 1-10  -> 0-9
    - 11-36 -> A-Z
    """
    if 1 <= remainder <= 10:
        # 1 映射为 '0', 10 映射为 '9'
        return str(remainder - 1)
    elif 11 <= remainder <= 36:
        # 11 映射为 'A', 36 映射为 'Z'
        # chr(65) 是 'A'
        return chr(remainder - 11 + 65)
    else:
        return '?'  # 理论上不应发生

def calculate_char_from_value(value):
    """
    对单个浮点数执行完整的AI码字符转换逻辑
    """
    # 1. 各位数相加
    sum_of_digits = sum_digits_from_float(value)
    
    # 2. 乘以 15
    result = sum_of_digits * 15
    
    # 3. 对 36 取余数
    remainder = result % 36
    
    # 4. 关键修正：余数 0 对应规则中的 36
    if remainder == 0:
        remainder = 36
        
    # 5. 映射到字符
    return map_remainder_to_char(remainder)

def get_data_driven_chars(data_array):
    """
    从 50x50 数据矩阵中提取10个点，计算并返回前后5位编码
    """
    # Excel 坐标 [J12, L14, ..., AB30]
    # 对应的 Python 索引 [行, 列]
    indices = [
        [11, 9],   # J12
        [13, 11],  # L14
        [15, 13],  # N16
        [17, 15],  # P18
        [19, 17],  # R20
        [21, 19],  # T22
        [23, 21],  # V24
        [25, 23],  # X26
        [27, 25],  # Z28
        [29, 27]   # AB30
    ]
    
    chars = []
    for (row, col) in indices:
        try:
            # 1. 从 50x50 数组中获取值
            value = data_array[row, col]
            
            # 2. 检查是否为 NaN (缺失数据)
            if np.isnan(value):
                chars.append('X')  # 如果数据缺失，用 'X' 填充
            else:
                # 3. 计算字符
                chars.append(calculate_char_from_value(value))
        except (IndexError, TypeError):
            chars.append('E')  # 如果索引越界或出错，用 'E' 填充

    # 4. 分割成前后缀
    prefix_code = "".join(chars[:5])
    suffix_code = "".join(chars[5:])
    
    return prefix_code, suffix_code

# 环曲 (TAC) 映射: { 数据库值: "编码" }
TAC_MAP = {
    0.00: '00',
    0.50: '01',  #  T1 = 0.50
    0.75: '02',  #  T2 = 0.75
    1.00: '03',  # T3 = 1.00
    1.25: '04',
    1.5: '05',
    1.75: '06',
    2.00: '07',
    2.25: '08',
    2.50: '09',
    2.75: '10',
    3.00: '11',
    3.25: '12',
    3.50: '13',
    3.75: '14',
    4.00: '15',
    4.25: '16',
    4.50: '17',
    4.75: '18',
    5.00: '19',
    5.25: '20',
    5.50: '21',
    5.75: '22',
    6.00: '23',
}

# 基弧曲率半径 映射: { 数据库值: "编码" }
BASE_ARC_RADIUS_MAP = {
    7.50: '01', 7.54: '02', 7.58: '03', 7.63: '04', 7.67: '05', 7.71: '06', 7.76: '07', 7.80: '08', 7.85: '09', 7.90: '10',
    7.94: '11', 7.99: '12', 8.04: '13', 8.08: '14', 8.13: '15', 8.18: '16', 8.23: '17', 8.28: '18', 8.33: '19', 8.39: '20',
    8.44: '21', 8.49: '22', 8.54: '23', 8.60: '24', 8.65: '25', 8.71: '26', 8.77: '27', 8.82: '28', 8.88: '29', 8.94: '30',
    9.00: '31', 9.06: '32', 9.12: '33', 9.18: '34', 9.25: '35', 9.31: '36', 9.38: '37', 9.44: '38', 9.51: '39', 9.57: '40',
    9.64: '41', 9.71: '42', 9.78: '43', 9.85: '44', 9.93: '45', 10.00: '46', 10.08: '47', 10.15: '48', 10.23: '49', 10.31: '50',
    10.39: '51', 10.47: '52', 10.55: '53', 10.63: '54', 10.71: '55',
}

# 配适弧K值 (AC K1) 映射: { 数据库值: "编码" }
AC_K_MAP = {
    40.00: '01', 40.25: '02', 40.50: '03', 40.75: '04', 41.00: '05', 41.25: '06', 41.50: '07', 41.75: '08', 42.00: '09',
    42.25: '10', 42.50: '11', 42.75: '12', 43.00: '13', 43.25: '14', 43.50: '15', 43.75: '16', 44.00: '17', 44.25: '18',
    44.50: '19', 44.75: '20', 45.00: '21', 45.25: '22', 45.50: '23', 45.75: '24', 46.00: '25', 46.25: '26', 46.50: '27',
    46.75: '28', 47.00: '29', 47.25: '30', 47.50: '31', 47.75: '32', 48.00: '33', 48.25: '34', 48.50: '35', 48.75: '36',
    49.00: '37', 37.00: '38', 37.25: '39', 37.50: '40', 37.75: '41', 38.00: '42', 38.25: '43', 38.50: '44', 38.75: '45',
    39.00: '46', 39.25: '47', 39.50: '48', 39.75: '49',
}

# 非球面系数 (ACe) 映射: { 数据库值: "编码" }
ACE_MAP = {
    0.00: '01',
    -0.25: '02',
    -0.50: '03',
    -0.75: '04',
    -1.00: '05',
}

# 总直径 映射: { 数据库值: "编码" }
DIAMETER_MAP = {
    9.50: 'A', 9.60: 'B', 9.70: 'C', 9.80: 'D', 9.90: 'E', 10.00: 'F', 10.10: 'G', 10.20: 'H',
    10.30: 'I', 10.40: 'J', 10.50: 'K', 10.60: 'L', 10.70: 'M', 10.80: 'N', 10.90: 'O', 11.00: 'P',
    11.10: 'Q', 11.20: 'R', 11.30: 'S', 11.40: 'T', 11.50: 'U', 11.60: 'V', 11.70: 'W', 11.80: 'X',
    11.90: 'Y', 12.00: 'Z',
}

# 光学区直径 映射: { 数据库值: "编码" }
OPTICAL_ZONE_MAP = {
    4.70: 'A', 4.80: 'B', 4.90: 'C', 5.00: 'D', 5.10: 'E', 5.20: 'F', 5.30: 'G', 5.40: 'H', 5.50: 'I',
    5.60: 'J', 5.70: 'K', 5.80: 'L', 5.90: 'M', 6.00: 'N', 6.10: 'O', 6.20: 'P', 6.30: 'Q', 6.40: 'R',
    6.50: 'S', 6.60: 'T', 6.70: 'U', 6.80: 'V', 6.90: 'W', 7.00: 'X',
}

# 配适弧径宽 (ac_arc_width) 映射: { 数据库值: "编码" }
ADAPTABLE_WIDTH_MAP = {
    0.40: 'A', 0.50: 'B', 0.60: 'C', 0.70: 'D', 0.80: 'E', 0.90: 'F', 1.00: 'G', 1.10: 'H',
    1.20: 'I', 1.30: 'J', 1.40: 'K', 1.50: 'L', 1.60: 'M', 1.70: 'N', 1.80: 'O', 1.90: 'P', 2.00: 'Q',
}

# 颜色 映射: { 数据库值: "编码" }
COLOR_MAP = {
    'left': 'B',  # 蓝色
    'right': 'G', # 绿色
}

# 边翘度 (Side Arc) 映射: { 数据库值: "编码" }
SIDE_ARC_MAP = {
    4.8: 'A',
    8.8: 'B',
    12.8: 'C',
    16.8: 'D',
}

# --- ↓↓↓ 新增映射 ↓↓↓ ---

# 光学区直径差异映射 (第27, 28位)
OPTICAL_DIFF_MAP = {
    -1.10: '00', -1.00: '01', -0.90: '02', -0.80: '03', -0.70: '04', -0.60: '05',
    -0.50: '06', -0.40: '07', -0.30: '08', -0.20: '09', -0.10: '10', 0.00: '11',
    0.10: '12', 0.20: '13', 0.30: '14', 0.40: '15', 0.50: '16', 0.60: '17',
    0.70: '18', 0.80: '19', 0.90: '20', 1.00: '21', 1.10: '22', 1.20: '23'
}

# 总直径计算差异映射 (第30位)
OVERALL_DIFF_MAP = {
    0.40: 'A', 0.50: 'B', 0.60: 'C', 0.70: 'D', 0.80: 'E', 0.90: 'F',
    1.00: 'G', 1.10: 'H', 1.20: 'I', 1.30: 'J', 1.40: 'K', 1.50: 'L',
    1.60: 'M', 1.70: 'N', 1.80: 'O', 1.90: 'P', 2.00: 'Q'
}

# --- 辅助函数 ---

def get_random_chars(length=5):
    """生成指定长度的随机大写字母和数字组合"""
    chars = string.ascii_uppercase + string.digits
    return ''.join(random.choice(chars) for _ in range(length))

def get_mapped_value(value_map, key, default='??'):
    """
    安全地从映射字典中获取值。
    能处理浮点数和 Decimal 类型的比较。
    """
    try:
        # 转换为 float 以便和字典的 float key 比较
        key_float = float(key)
    except (ValueError, TypeError, SystemError):
        return default # 处理 None, 空字符串, 或无效值

    # 优先尝试精确匹配 (适用于 0.0, 1.0 等)
    if key_float in value_map:
        return value_map[key_float]

    # 尝试模糊匹配 (处理 7.500001 和 7.50 的问题)
    for k_map, v_map in value_map.items():
        if np.isclose(key_float, k_map):
            return v_map
            
    return default # 找不到匹配项


def generate_ai_code(params, custom, data_array=None):
    """根据 BasicParams 和 ACCustomization 实例生成 AI 码"""
    if not custom:
        return "" # 如果没有定制数据，返回空

    try:
        # 1. 环曲 (TAC)
        tac_code = get_mapped_value(TAC_MAP, custom.tac_position, default='~')
        
        # 2. 基弧曲率半径
        rounded_radius = round(float(custom.base_arc_curvature_radius), 2)
        radius_code = get_mapped_value(BASE_ARC_RADIUS_MAP, rounded_radius, default='!')

        # 3. 配适弧K值
        ac_k_code = get_mapped_value(AC_K_MAP, custom.ac_arc_k1, default='@')

        # 4. 非球面系数 (ACe)
        ace_code = get_mapped_value(ACE_MAP, custom.ace_position, default='#')

        # 5. 总直径
        diameter_code = get_mapped_value(DIAMETER_MAP, params.overall_diameter, default='￥')

        # 6. 光学区直径 (第11位)
        optical_code = get_mapped_value(OPTICAL_ZONE_MAP, params.optical_zone_diameter, default='%')

        # 7. 配适弧径宽 (第12位)
        adapt_width_code = get_mapped_value(ADAPTABLE_WIDTH_MAP, params.ac_arc_width, default='^')

        # 8. 颜色
        color_code = COLOR_MAP.get(params.eye, '*')

        # 9. 边翘度
        side_arc_code = get_mapped_value(SIDE_ARC_MAP, custom.side_arc_position, default='?')

        # --- ↓↓↓ 新增计算逻辑 ↓↓↓ ---
        
        # 准备数据
        optical_zone_val = float(params.optical_zone_diameter)
        overall_diameter_val = float(params.overall_diameter)

        # 第27、28位：光学区直径 - 5.8
        optical_diff = round(optical_zone_val - 5.8, 2)
        code_27_28 = get_mapped_value(OPTICAL_DIFF_MAP, optical_diff, default='00')

        # 第29位：固定位 L
        code_29 = 'L'

        # 第30位：(总直径 - 8.4) / 2
        overall_diff = round((overall_diameter_val - 8.4) / 2, 2)
        code_30 = get_mapped_value(OVERALL_DIFF_MAP, overall_diff, default='A')

        # 判断并调换位置
        # 如果光学区直径小于 5.8
        if optical_zone_val < 5.8:
            # 交换 第29位(code_29) 与 第11位(optical_code)
            temp_29 = code_29
            code_29 = optical_code
            optical_code = temp_29

            # 交换 第30位(code_30) 与 第12位(adapt_width_code)
            temp_30 = code_30
            code_30 = adapt_width_code
            adapt_width_code = temp_30

        # --- ↑↑↑ 新增计算逻辑结束 ↑↑↑ ---

        # 组装核心编码 (中间部分)
        parts = [
            'W',                     # 1. 固定 'W'
            tac_code,                # 2. 环曲
            radius_code,             # 3. 基弧
            ac_k_code,               # 4. 配适弧K
            ace_code,                # 5. 非球面系数
            diameter_code,           # 6. 总直径
            optical_code,            # 7. 光学区 (可能已交换)
            adapt_width_code,        # 8. 配适弧径宽 (可能已交换)
            # reverse_width_code,    # 9. 反转弧径宽 (暂无)
            color_code,              # 10. 颜色
            '22',                    # 11. 固定 '22'
            side_arc_code            # 12. 边翘度
        ]
        
        core_code = "".join(parts)
        
        # 检查是否传入了 50x50 数据矩阵 (仅 Medmont 会传入)
        if data_array is not None and (params.custom_type == "0" or params.custom_type == "2"):
            # 使用新逻辑从10个数据点生成编码 (前5位，后5位)
            prefix_code, suffix_code = get_data_driven_chars(data_array)
            print("使用数据驱动逻辑生成前后缀")
        else:
            # 否则，回退到旧的随机码逻辑
            prefix_code = get_random_chars(5)
            suffix_code = get_random_chars(5)
            print("使用随机码逻辑生成前后缀")
            
        # 返回最终编码 (30位)
        # 前缀(5) + 核心(16) + 后缀(5) + 新增(4) = 30
        return f"{prefix_code}{core_code}{suffix_code}{code_27_28}{code_29}{code_30}"

    except Exception as e:
        # logger.error(f"生成AI码失败 (PatientID: {params.patient.id}, Eye: {params.eye}): {e}")
        print(f"生成AI码失败: {e}")
        return "GENERATION_ERROR"
# ==========================================================
# =============   ↑↑↑ 2. AI码 生成逻辑结束 ↑↑↑   ============

class ReviewResultUpdateView(UpdateView):
    model = ReviewResult
    form_class = ReviewResultForm
    template_name = "patient/../templates/uppatient/medment_index_4_customize.html"
    success_url = "/success/"


# 导出定制数据
class ExportCustomizedView(View):
    template_name = 'upexportimport/export_customized.html'

    def get(self, request, *args, **kwargs):
        # 只显示当前用户创建的患者
        patients = Patient.objects.filter(created_by=request.user).order_by('-create_date')
        return render(request, self.template_name, {'patients': patients})

    def post(self, request, *args, **kwargs):
        patient_ids = request.POST.getlist('patient_ids')
        if not patient_ids:
            messages.error(request, "您没有选择任何要导出的患者。")
            return redirect('export_customized')

        response = HttpResponse(
            content_type='text/csv',
            headers={'Content-Disposition': 'attachment; filename="full_patient_data_export.csv"'},
        )
        response.write(u'\ufeff'.encode('utf8'))
        writer = csv.writer(response)

        # ==========================================================
        # =============   ↓↓↓ 1. 定义完整的表头 ↓↓↓   ================
        # ==========================================================
        headers = [
            '终端客户', '患者ID', '姓名','测量设备', '眼别', '基弧曲率半径', '光学区直径',
            '反转弧径宽', '配适弧径宽', '总直径', '颜色', 'AI码', '备注',
            '是否加急费', '市场活动'
        ]
        writer.writerow(headers)
        # ==========================================================
        
        # 2. 根据勾选的患者ID，查询所有相关的 BasicParams 数据
        # prefetch_related 会一次性高效加载所有关联数据
        params_to_export = BasicParams.objects.filter(
            patient_id__in=patient_ids
        ).select_related(
            'patient'
        ).prefetch_related(
            'cornealtopography_set',
            'accustomization_set'
        ).order_by('patient__name', 'eye')

        # 3. 遍历数据并写入 CSV
        for params in params_to_export:
            # 安全地获取关联对象
            topo = params.cornealtopography_set.first()
            custom = params.accustomization_set.first()
            
            # --- ↓↓↓ 新增:根据 custom_type 确定设备名称 ↓↓↓ ---
            device_name = params.get_custom_type_display() 
            if params.custom_type == '0':
                device_name = 'Medmont'
            elif params.custom_type == '4':
                device_name = 'Tomey'
            elif params.custom_type == '2':
                device_name = 'Suore'
            # --- ↓↓↓ 3. 从数据库读取已保存的 AI 码 ↓↓↓ ---
            ai_code = custom.ai_code if custom and custom.ai_code else '' # <-- 直接读取字段


            writer.writerow([
                request.user.username,  # 终端客户（当前登录用户名）
                params.patient.id,  # 患者ID
                params.patient.name,  # 姓名
                # params.get_custom_type_display(),  # 测量设备
                device_name,  # 测量设备
                params.get_eye_display(),  # 眼别
                custom.base_arc_curvature_radius if custom else '',  # 基弧曲率半径
                params.optical_zone_diameter,  # 光学区直径
                custom.reverse_arc_width if custom else '',  # 反转弧径宽
                params.ac_arc_width if custom else '', # 配适弧径宽
                params.overall_diameter,  # 总直径
                '蓝色' if params.eye == 'left' else '绿色',  # 颜色
                ai_code,  # AI码
                '',  # 备注 (留空)
                '',  # 是否加急费 (留空)
                '',  # 市场活动 (留空)
            ])

            
        return response

class GenerateExportDataView(View):
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            patient_ids = request.POST.getlist('patient_ids')
            # 从 POST 参数获取用户名（由前端传递）
            username = request.POST.get('username', '')
            if not patient_ids:
                 return JsonResponse({'success': False, 'error': '未提供患者ID。'}, status=400)

            # 创建 Excel 工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "导出数据"

            # 定义表头
            headers = [
                '终端客户', '姓名', '眼别', '基弧曲率半径', '光学区直径',
                '反转弧径宽', '配适弧径宽', '总直径', '颜色', 'AI码', '备注',
                '是否加急费', '市场活动'
            ]
            ws.append(headers)

            # 查询数据
            params_to_export = BasicParams.objects.filter(
                patient_id__in=patient_ids
            ).select_related(
                'patient'
            ).prefetch_related(
                'cornealtopography_set',
                'accustomization_set'
            ).order_by('patient__name', 'eye')

            # 按患者分组，记录每个患者的起始行和结束行
            current_row = 2  # 从第2行开始（第1行是表头）
            patient_rows = {}  # {patient_id: [start_row, end_row]}

            for params in params_to_export:
                topo = params.cornealtopography_set.first()
                custom = params.accustomization_set.first()

                ai_code = custom.ai_code if custom and custom.ai_code else ''

                row_data = [
                    username,  # 终端客户（从前端传递的用户名）
                    params.patient.name,  # 姓名
                    params.get_eye_display(),  # 眼别
                    custom.base_arc_curvature_radius if custom else '',  # 基弧曲率半径
                    params.optical_zone_diameter,  # 光学区直径
                    custom.reverse_arc_width if custom else '',  # 反转弧径宽
                    params.ac_arc_width if custom else '',  # 配适弧径宽
                    params.overall_diameter,  # 总直径
                    '蓝色' if params.eye == 'left' else '绿色',  # 颜色
                    ai_code,  # AI码
                    '',  # 备注
                    '',  # 是否加急费
                    '',  # 市场活动
                ]
                ws.append(row_data)

                # 记录患者行范围
                patient_id = params.patient.id
                if patient_id not in patient_rows:
                    patient_rows[patient_id] = [current_row, current_row]
                else:
                    patient_rows[patient_id][1] = current_row

                current_row += 1

            # 合并终端客户单元格（A列）和姓名单元格（B列）
            for patient_id, (start_row, end_row) in patient_rows.items():
                if end_row > start_row:  # 只有多于一行时才合并
                    ws.merge_cells(f'A{start_row}:A{end_row}')
                    ws[f'A{start_row}'].alignment = Alignment(vertical='center', horizontal='center')
                    ws.merge_cells(f'B{start_row}:B{end_row}')
                    ws[f'B{start_row}'].alignment = Alignment(vertical='center', horizontal='center')

            # 将 Excel 文件保存到内存
            excel_io = io.BytesIO()
            wb.save(excel_io)
            excel_io.seek(0)

            # 转换为 base64 编码
            excel_base64 = base64.b64encode(excel_io.getvalue()).decode('utf-8')

            return JsonResponse({'success': True, 'excel_data': excel_base64})

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class UpdateExportCountView(View):
    """
    更新患者导出次数的 API
    成功导出后调用此接口增加导出次数
    """
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            patient_ids = request.POST.getlist('patient_ids')
            if not patient_ids:
                return JsonResponse({'success': False, 'error': '未提供患者ID。'}, status=400)

            # 更新每个患者的导出次数
            from django.db.models import F
            updated_count = Patient.objects.filter(id__in=patient_ids).update(
                export_count=F('export_count') + 1
            )

            return JsonResponse({
                'success': True,
                'message': f'成功更新 {updated_count} 个患者的导出次数'
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class ImportCustomizedView(TemplateView):
    # template_name = 'patient/import_customized.html'
    # template_name = 'upexportimport/import_customized.html'
    template_name = 'upexceptionpage/404.html'

    def get(self, request, *args, **kwargs):
        try:
            return super().get(request, *args, **kwargs)
        except Exception as e:
            return redirect('error_page')


class ExportReviewedView(TemplateView):
    # template_name = 'patient/export_reviewed.html'
    # template_name = 'upexportimport/export_reviewed.html'
    template_name = 'upexceptionpage/404.html'

    def get(self, request, *args, **kwargs):
        try:
            return super().get(request, *args, **kwargs)
        except Exception as e:
            return redirect('error_page')


class ImportReviewedView(TemplateView):
    template_name = 'upexceptionpage/404.html'

    def get(self, request, *args, **kwargs):
        try:
            return super().get(request, *args, **kwargs)
        except Exception as e:
            return redirect('error_page')


class ParamsModifyView(UpdateView):
    model = ACCustomization
    template_name = 'memdent/paras_modify.html'
    form_class = ParamsModifyForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # 基本信息
        ac_customization = ACCustomization.objects.filter(id=self.object.pk).last()
        basic_params_data = ac_customization.BasicParams
        relationship = RelationshipTable.objects.filter(lens_type=basic_params_data.lens_type, belongs_level=1).values()

        context['ac_customization'] = ac_customization
        context['basic_params_data'] = basic_params_data
        context["relationship"] = relationship
        context['choice_options'] = {
            "tac_options": tac_options,
            "ac_arc_options": ac_arc_options,
            "ace_position_options": ace_position_options,
            "side_arc_position_options": side_arc_position_options,
            "reverse_arc_height_options": reverse_arc_height_options,
        }
        return context

    def form_invalid(self, form):
        # 打印错误信息到控制台
        # print(form.errors)
        messages.error(self.request, '表单验证失败，请检查输入')
        return super().form_invalid(form)

    # def form_valid(self, form):
    #     print("表单数据:", form.cleaned_data)
    #     ac_pk = self.object.pk
    #     if form.is_valid():
    #         form.save()
    #         updated = self.secondary_caluculation(ac_pk)
    #         if updated:
    #             messages.success(self.request, '参数更新成功')
    #             success_url = reverse('params_modify', kwargs={'pk': ac_pk})
    #             return HttpResponseRedirect(success_url)
    #         else:
    #             messages.error(self.request, '参数更新失败')
    #             # 添加返回语句，确保用户能留在当前页面
    #             return self.form_invalid(form)

    def form_valid(self, form):
        # Django 的 UpdateView 会自动处理 form.save()
        # super().form_valid(form) 会保存表单并返回重定向响应
        # 但我们需要在重定向之前执行二次计算，所以我们自己处理
        
        if not form.is_valid():
            return self.form_invalid(form)

        # 1. 保存用户在表单中所做的修改
        # self.object 会被更新为保存后的实例
        self.object = form.save()
        
        # 2. 调用二次计算
        updated = self.secondary_caluculation(self.object.pk)
        
        if updated:
            messages.success(self.request, '参数替换成功！')
        else:
            # 即使二次计算失败，主参数也已更新，可以只给一个警告
            messages.warning(self.request, '参数已更新，但二次计算失败。')

        # 3. 获取患者ID并构建重定向URL
        patient_pk = self.object.patient.pk
        success_url = reverse('patient_detail', kwargs={'pk': patient_pk})
        
        return HttpResponseRedirect(success_url)

    @staticmethod
    def secondary_caluculation(ac_pk):
        ac_customization_data = ACCustomization.objects.filter(id=ac_pk).last()
        basic_params_data = ac_customization_data.BasicParams
        cornealtopography_data = CornealTopography.objects.filter(BasicParams=basic_params_data).last()
        # 更新泪膜图数据
        data_file = basic_params_data.corneal_file
        full_path = os.path.join(settings.MEDIA_ROOT, str(data_file))
        full_path = os.path.normpath(full_path)

        common_params = {
            'lens_type': basic_params_data.lens_type,
            'ace_position': float(ac_customization_data.ace_position),
            'flat_k': float(cornealtopography_data.flat_k),
            'base_arc_curvature_radius': float(ac_customization_data.base_arc_curvature_radius),
            'side_arc_position': float(ac_customization_data.side_arc_position),
            'ac_arc_start': float(basic_params_data.ac_arc_start),
            'ac_arc_end': float(basic_params_data.ac_arc_end),
            'reverse_arc_height': float(ac_customization_data.reverse_arc_height),
            'overall_diameter': float(basic_params_data.overall_diameter),
            'al_type': basic_params_data.custom_type,
            'file_path': full_path
        }

        # print(f"文件位置:{full_path}")
        if basic_params_data.custom_type == "0" or basic_params_data.custom_type == "2":

            # 计算平K泪膜图
            tear_film_ping_k = TearFilmHeightCalculator(ac_arc_k1=float(ac_customization_data.ac_arc_k1),
                                                        degree_list=[float(cornealtopography_data.plane_angle),
                                                                     float(cornealtopography_data.plane_angle + 180)],
                                                        **common_params).main_calculate()
            C = TEARFILMDATA(tear_film_ping_k['x'], tear_film_ping_k['y1'], tear_film_ping_k['y2'])
            tear_film_ping_k_data = C.main()

            # 计算陡K泪膜图
            tear_film_steep_k = TearFilmHeightCalculator(ac_arc_k1=float(ac_customization_data.steep_k_calculate),
                                                         degree_list=[float(cornealtopography_data.inclined_angle),
                                                                      float(
                                                                          cornealtopography_data.inclined_angle + 180)],
                                                         **common_params).main_calculate()
            C = TEARFILMDATA(tear_film_steep_k['x'], tear_film_steep_k['y1'], tear_film_steep_k['y2'])
            tear_film_steep_k_data = C.main()

            # 保存泪膜图
            tear_film = {
                "tear_film_ping_k": tear_film_ping_k_data,
                "ping_k_lens_height": tear_film_ping_k['lens_height'],
                "ping_k_radius_list01": tear_film_ping_k['radius_list01'],
                "ping_k_radius_list02": tear_film_ping_k['radius_list02'],
                "tear_film_steep_k": tear_film_steep_k_data,
                "steep_k_lens_height": tear_film_steep_k['lens_height'],
                "steep_k_radius_list01": tear_film_steep_k['radius_list01'],
                "steep_k_radius_list02": tear_film_steep_k['radius_list02'],
                "bc": basic_params_data.ac_arc_start - 0.8,
                "rc": basic_params_data.ac_arc_start,
                "ac": basic_params_data.ac_arc_end,
                "pc": basic_params_data.overall_diameter / 2,
            }

            print(basic_params_data)
            # 生成二维码
            # 注意：此处二维码逻辑生成 AI 码需要调用 generate_ai_code，这里暂时只更新了 tear_film
            # 如果需要二维码中包含新AI码，应先重新生成AI码。根据现有逻辑，AI码是在 BasicParams/ACCustomization 
            # 创建时或修改时生成的。此处仅更新了 tear_film_data。
            
            # 重新生成AI码并保存
            new_ai_code = generate_ai_code(basic_params_data, ac_customization_data, data_array=None) # 注意：这里data_array可能需要传入，但在二次计算中可能获取不到原始数据矩阵，如果不传会用随机码填充前后缀，需根据业务确认

            qrcode_data = f"患者ID:{basic_params_data.patient_id}\n角膜地形图ID:{cornealtopography_data.id}\n定制参数如下:\n" \
                          f"AC弧K值(D):{ac_customization_data.ac_arc_k1}\n" \
                          f"基弧曲率半径(mm):{ac_customization_data.base_arc_curvature_radius}\n" \
                          f"Tac档位:{ac_customization_data.tac_position}\n" \
                          f"ACe档位:{ac_customization_data.ace_position}\n" \
                          f"反转弧矢高:{ac_customization_data.reverse_arc_height}\n" \

            qrcode_path = txt_to_qrcode(qrcode_data).get("save_dir_path", None)


            # 保存结果
            updated = ACCustomization.objects.filter(id=ac_customization_data.id).update(
                tear_film_data=tear_film,
                qrcode_medment_accustomization=qrcode_path,
                ai_code=new_ai_code # 更新 AI 码
            )

            # 荧光染色图
            fluorescent_staining = FluorescentStaining(degree_list=[float(cornealtopography_data.plane_angle),
                                                                    float(cornealtopography_data.plane_angle + 180),
                                                                    float(cornealtopography_data.inclined_angle),
                                                                    float(cornealtopography_data.inclined_angle) + 180],
                                                       acc_id=ac_customization_data.id,
                                                       overall_diameter=basic_params_data.overall_diameter,
                                                       optical_zone_diameter=basic_params_data.optical_zone_diameter, )
            result = fluorescent_staining.fluorescent_staining(
                h1=tear_film_ping_k['y1'],
                h3=tear_film_ping_k['y2'],
                h2=tear_film_steep_k['y1'],
                h4=tear_film_steep_k['y2']
            )
            updated = result["updated"]
            return updated
        elif basic_params_data.custom_type == "1" or basic_params_data.custom_type == "3":
            # 生成泪膜图数据

            # print(f"common_params:{common_params}")

            # 计算平K泪膜图
            degree_list = [int(cornealtopography_data.plane_angle),
                           int(cornealtopography_data.plane_angle) + 90,
                           int(cornealtopography_data.plane_angle) + 180,
                           int(cornealtopography_data.plane_angle) + 270]
            # print(f"degree_list:{degree_list}")
            # print(f"common_params:{common_params}")
            # print(ac_customization_data.ac_arc_k1)
            # print(ac_customization_data.ac_arc_k2)
            # print(ac_customization_data.ac_arc_k3)
            # print(ac_customization_data.ac_arc_k4)
            tear_film_ping_k = TearFilmHeightCalculator(ac_arc_k1=float(ac_customization_data.ac_arc_k1),
                                                        ac_arc_k2=float(ac_customization_data.ac_arc_k2),
                                                        ac_arc_k3=float(ac_customization_data.ac_arc_k3),
                                                        ac_arc_k4=float(ac_customization_data.ac_arc_k4),
                                                        degree_list=degree_list,
                                                        **common_params).main_calculate()
            # 计算平K泪膜图
            C1 = TEARFILMDATA(tear_film_ping_k['x'], tear_film_ping_k['y1'], tear_film_ping_k['y3'])
            tear_film_ping_k_data_01 = C1.main()
            C2 = TEARFILMDATA(tear_film_ping_k['x'], tear_film_ping_k['y2'], tear_film_ping_k['y4'])
            tear_film_ping_k_data_02 = C2.main()

            # 保存泪膜图
            tear_film = {
                "tear_film_ping_k": tear_film_ping_k_data_01,
                "tear_film_steep_k": tear_film_ping_k_data_02,
                "ping_k_lens_height_01": tear_film_ping_k['lens_height_01'],
                "ping_k_lens_height_02": tear_film_ping_k['lens_height_02'],
                "ping_k_lens_height_03": tear_film_ping_k['lens_height_03'],
                "ping_k_lens_height_04": tear_film_ping_k['lens_height_04'],
                "ping_k_radius_list01": tear_film_ping_k['radius_list01'],
                "ping_k_radius_list02": tear_film_ping_k['radius_list02'],
                "ping_k_radius_list03": tear_film_ping_k['radius_list03'],
                "ping_k_radius_list04": tear_film_ping_k['radius_list04'],
                "bc": basic_params_data.ac_arc_start - 0.8,
                "rc": basic_params_data.ac_arc_start,
                "ac": basic_params_data.ac_arc_end,
                "pc": basic_params_data.overall_diameter / 2,
            }
            
            # 重新生成 AI 码
            new_ai_code = generate_ai_code(basic_params_data, ac_customization_data, data_array=None)

            # 保存结果
            updated = ACCustomization.objects.filter(id=ac_customization_data.id).update(
                tear_film_data=tear_film,
                base_arc_curvature_radius=tear_film_ping_k["min_base_arc_curvature_radius"],
                ai_code=new_ai_code
            )
            return updated
        elif basic_params_data.custom_type == "4":
            # 生成泪膜图数据
            rm_file_path = os.path.join(settings.MEDIA_ROOT, str(basic_params_data.corneal_file))
            rm_file_path = os.path.normpath(rm_file_path)
            ch_file_path = os.path.join(settings.MEDIA_ROOT, str(basic_params_data.corneal_file2))
            ch_file_path = os.path.normpath(ch_file_path)

            common_params["rm_file"] = rm_file_path # 半径文件
            common_params["ch_file"] = ch_file_path # 高度文件

            # 计算平K泪膜图
            tear_film_ping_k = TearFilmHeightCalculator(ac_arc_k1=float(ac_customization_data.ac_arc_k1),
                                                        degree_list=[float(cornealtopography_data.plane_angle),
                                                                     float(cornealtopography_data.plane_angle + 180)],
                                                        **common_params).main_calculate()
            C = TEARFILMDATA(tear_film_ping_k['x'], tear_film_ping_k['y1'], tear_film_ping_k['y2'])
            tear_film_ping_k_data = C.main()

            # 计算陡K泪膜图
            tear_film_steep_k = TearFilmHeightCalculator(ac_arc_k1=float(ac_customization_data.steep_k_calculate),
                                                         degree_list=[float(cornealtopography_data.inclined_angle),
                                                                      float(
                                                                          cornealtopography_data.inclined_angle + 180)],
                                                         **common_params).main_calculate()
            C = TEARFILMDATA(tear_film_steep_k['x'], tear_film_steep_k['y1'], tear_film_steep_k['y2'])
            tear_film_steep_k_data = C.main()

            # 保存泪膜图
            tear_film = {
                "tear_film_ping_k": tear_film_ping_k_data,
                "ping_k_lens_height": tear_film_ping_k['lens_height'],
                "ping_k_radius_list01": tear_film_ping_k['radius_list01'],
                "ping_k_radius_list02": tear_film_ping_k['radius_list02'],
                "tear_film_steep_k": tear_film_steep_k_data,
                "steep_k_lens_height": tear_film_steep_k['lens_height'],
                "steep_k_radius_list01": tear_film_steep_k['radius_list01'],
                "steep_k_radius_list02": tear_film_steep_k['radius_list02'],
                "bc": basic_params_data.ac_arc_start - 0.8,
                "rc": basic_params_data.ac_arc_start,
                "ac": basic_params_data.ac_arc_end,
                "pc": basic_params_data.overall_diameter / 2,
            }
            
            # 重新生成 AI 码
            new_ai_code = generate_ai_code(basic_params_data, ac_customization_data, data_array=None)

            # 保存结果
            updated = ACCustomization.objects.filter(id=ac_customization_data.id).update(
                tear_film_data=tear_film,
                ai_code=new_ai_code
            )

            # 荧光染色图
            fluorescent_staining = FluorescentStaining(degree_list=[float(cornealtopography_data.plane_angle),
                                                                    float(cornealtopography_data.plane_angle + 180),
                                                                    float(cornealtopography_data.inclined_angle),
                                                                    float(cornealtopography_data.inclined_angle) + 180],
                                                       acc_id=ac_customization_data.id,
                                                       overall_diameter=basic_params_data.overall_diameter,
                                                       optical_zone_diameter=basic_params_data.optical_zone_diameter, )
            result = fluorescent_staining.fluorescent_staining(
                h1=tear_film_ping_k['y1'],
                h3=tear_film_ping_k['y2'],
                h2=tear_film_steep_k['y1'],
                h4=tear_film_steep_k['y2']
            )
            updated = result["updated"]
            return updated
        elif basic_params_data.custom_type == "5":
            # 生成泪膜图数据
            rm_file_path = os.path.join(settings.MEDIA_ROOT, str(basic_params_data.corneal_file))
            rm_file_path = os.path.normpath(rm_file_path)
            ch_file_path = os.path.join(settings.MEDIA_ROOT, str(basic_params_data.corneal_file2))
            ch_file_path = os.path.normpath(ch_file_path)

            common_params["rm_file"] = rm_file_path  # 半径文件
            common_params["ch_file"] = ch_file_path  # 高度文件

            # 计算平K泪膜图
            degree_list = [float(cornealtopography_data.plane_angle),
                           float(cornealtopography_data.plane_angle) + 90,
                           float(cornealtopography_data.plane_angle) + 180,
                           float(cornealtopography_data.plane_angle) + 270]
            # print(f"degree_list:{degree_list}")
            tear_film_ping_k = TearFilmHeightCalculator(ac_arc_k1=float(ac_customization_data.ac_arc_k1),
                                                        ac_arc_k2=float(ac_customization_data.ac_arc_k2),
                                                        ac_arc_k3=float(ac_customization_data.ac_arc_k3),
                                                        ac_arc_k4=float(ac_customization_data.ac_arc_k4),
                                                        degree_list=degree_list,
                                                        **common_params).main_calculate()
            # 计算平K泪膜图
            C1 = TEARFILMDATA(tear_film_ping_k['x'], tear_film_ping_k['y1'], tear_film_ping_k['y3'])
            tear_film_ping_k_data_01 = C1.main()
            C2 = TEARFILMDATA(tear_film_ping_k['x'], tear_film_ping_k['y2'], tear_film_ping_k['y4'])
            tear_film_ping_k_data_02 = C2.main()

            # 保存泪膜图
            tear_film = {
                "tear_film_ping_k": tear_film_ping_k_data_01,
                "tear_film_steep_k": tear_film_ping_k_data_02,
                "ping_k_lens_height_01": tear_film_ping_k['lens_height_01'],
                "ping_k_lens_height_02": tear_film_ping_k['lens_height_02'],
                "ping_k_lens_height_03": tear_film_ping_k['lens_height_03'],
                "ping_k_lens_height_04": tear_film_ping_k['lens_height_04'],
                "ping_k_radius_list01": tear_film_ping_k['radius_list01'],
                "ping_k_radius_list02": tear_film_ping_k['radius_list02'],
                "ping_k_radius_list03": tear_film_ping_k['radius_list03'],
                "ping_k_radius_list04": tear_film_ping_k['radius_list04'],
                "bc": basic_params_data.ac_arc_start - 0.8,
                "rc": basic_params_data.ac_arc_start,
                "ac": basic_params_data.ac_arc_end,
                "pc": basic_params_data.overall_diameter / 2,
            }
            
            # 重新生成 AI 码
            new_ai_code = generate_ai_code(basic_params_data, ac_customization_data, data_array=None)

            # 保存结果
            updated = ACCustomization.objects.filter(id=ac_customization_data.id).update(
                tear_film_data=tear_film,
                base_arc_curvature_radius=tear_film_ping_k["min_base_arc_curvature_radius"],
                ai_code=new_ai_code
            )
            return updated



class ErrorView(TemplateView):
    template_name = 'upexceptionpage/error.html'